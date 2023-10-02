from openpyxl import load_workbook
import smtplib

from email.mime.text import MIMEText  # 邮件正文
from email.header import Header  # 邮件头

# 加载excel文件，data_only=True(excel中工资为函数计算的值，邮件中要把公式转换为计算值)
wb = load_workbook("5月工资.xlsx", data_only=True)

sheet = wb.active


# 登录网易邮箱
smtp_obj = smtplib.SMTP_SSL("smtp.163.com", 465)
smtp_obj.login("xxx@163.com", "xxxxxx")

# 循环excel
count = 0
table_col_html = "<thread>"  # 表格表头
for row in sheet.iter_rows(min_row=1, max_row=26):
    count += 1
    if count == 1:  # first row
        for col in row:
            table_col_html += f"<th>{col.value}</th>"
        table_col_html += "</thread>"
        continue
    else:
        row_text = "<tr>"  # 表格开始一行
        for cell in row:
            row_text += f"<td>{cell.value}</td>"
        row_text += "</tr>"  # 表格结束一行
        name = row[2]
        staff_email = row[1].value
        print(staff_email, name.value)

    # 邮件内容
    mail_body_context = f"""
        <h3>{name.value},你好:</h3>
        <p>请查收你的工资条。。。。</p>
        <table border="1px solid black"> 
          {table_col_html}
         {row_text}
       </table>
    """
    msg_body = MIMEText(mail_body_context, "html", "utf-8")

    # 判断每发送10封重新登录一次（本脚本测试得知网易邮箱单次登录最多发送11封邮件，并且不适用于qq邮箱 qq邮箱会检测为垃圾邮件）
    if count % 10 == 0:
        smtp_obj = smtplib.SMTP_SSL("smtp.163.com", 465)
        smtp_obj.login("xxx@163.com", "xxxxxx")

    msg_body["From"] = Header("人事部", "utf-8")  # 发送者
    msg_body["To"] = Header("员工", "utf-8")  # 接收者
    msg_body["Subject"] = Header("2021年5月工资", "utf-8")  # 主题

    # 发邮件
    smtp_obj.sendmail(
        "xxx@163.com",
        [
            staff_email,
        ],
        msg_body.as_string(),
    )
    print(f"成功发送工资条到{staff_email}- {name.value}....")
